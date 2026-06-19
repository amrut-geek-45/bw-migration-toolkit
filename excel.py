import streamlit as st
import pandas as pd
import os
import glob
import re
import io
import zipfile
from collections import Counter, defaultdict

from analyzer import analyze_code
from transformation_code import process_transformation

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="BW Transformation Analyzer",
    page_icon="⚙️",
    layout="wide"
)

# =====================================================
# CSS
# =====================================================

st.markdown("""
<style>
.block-container{padding-top:2rem;padding-bottom:2rem}
.hero-title{font-size:40px;font-weight:700;color:white;margin-bottom:4px}
.hero-sub{font-size:16px;color:#A0A0A0;margin-bottom:1.5rem}
.metric-card{background:#161B22;padding:20px;border-radius:12px;
             border:1px solid #30363D;text-align:center;color:white}
.metric-card h3{font-size:12px;color:#8b949e;margin-bottom:6px;
                text-transform:uppercase;letter-spacing:.06em}
.metric-card h2{font-size:28px;font-weight:700;margin:0}
.cx-LOW   {color:#3fb950}
.cx-MEDIUM{color:#e3b341}
.cx-HIGH  {color:#f85149}
.ai-box{background:#0d1117;border-left:4px solid #6f42c1;padding:14px 18px;
        border-radius:6px;color:#c9d1d9;font-size:14px;line-height:1.7;margin:8px 0}
.ztable-card{background:#0d1117;border:1px solid #30363d;border-radius:8px;
             padding:10px 14px;font-size:13px;color:#c9d1d9;margin:4px 0}
.ln-wrap{font-family:monospace;margin:2px 0}
.ln-row{display:flex;align-items:center;gap:6px;padding:6px 12px;
        border-radius:6px;margin:2px 0;font-size:13px;color:#c9d1d9;
        white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ln-query      {background:#1f1b2e;border-left:4px solid #a78bfa}
.ln-provider   {background:#162032;border-left:4px solid #38bdf8}
.ln-mp         {background:#162032;border-left:4px solid #1a6ea8}
.ln-cube       {background:#0f2a1e;border-left:4px solid #3fb950}
.ln-adso       {background:#271d0f;border-left:4px solid #e3b341}
.ln-trfn       {background:#1e1120;border-left:4px solid #bc8cff}
.ln-dtp        {background:#0f2020;border-left:4px solid #2dd4bf}
.ln-datasource {background:#200f0f;border-left:4px solid #f85149}
.ln-infosource {background:#0f1f2e;border-left:4px solid #38bdf8}
.ln-infopackage{background:#1a0f2e;border-left:4px solid #a78bfa}
.ln-srcsys     {background:#1a1a0f;border-left:4px solid #fbbf24}
.ln-default    {background:#161b22;border-left:4px solid #30363d}
.ln-connector  {opacity:.35;margin-right:2px;user-select:none}
.ln-type       {font-size:10px;opacity:.6;font-weight:700;
                text-transform:uppercase;letter-spacing:.05em;flex-shrink:0}
.ln-name       {font-size:12px;background:rgba(255,255,255,.06);
                padding:1px 6px;border-radius:3px;overflow:hidden;
                text-overflow:ellipsis}
.ln-legend     {display:flex;flex-wrap:wrap;gap:6px;margin-top:10px}
.ln-leg-item   {display:flex;align-items:center;gap:5px;
                padding:3px 10px;border-radius:20px;font-size:11px;color:#c9d1d9}
div[data-testid="metric-container"]{background:#161b22;border:1px solid #30363d;
    border-radius:10px;padding:12px 16px}
div[data-testid="stDownloadButton"] button{background:#238636;color:white;
    border:none;border-radius:6px;font-weight:600;padding:8px 20px}
div[data-testid="stDownloadButton"] button:hover{background:#2ea043}
.ai-sum-box{background:#0d1117;border-left:4px solid #6f42c1;padding:14px 18px;
    border-radius:6px;color:#c9d1d9;font-size:14px;line-height:1.7;margin:8px 0}
.risk-CRITICAL{background:#3d0a0a;border-left:5px solid #f85149;padding:12px 16px;
    border-radius:6px;font-size:13px;margin:6px 0;color:#ffa198}
.risk-HIGH{background:#2d1a00;border-left:5px solid #e3b341;padding:12px 16px;
    border-radius:6px;font-size:13px;margin:6px 0;color:#f0c674}
.risk-MEDIUM{background:#1a2200;border-left:5px solid #7ee787;padding:12px 16px;
    border-radius:6px;font-size:13px;margin:6px 0;color:#9be89e}
.risk-LOW{background:#0d1f2d;border-left:5px solid #38bdf8;padding:12px 16px;
    border-radius:6px;font-size:13px;margin:6px 0;color:#7dd3fc}
.risk-banner-CRITICAL{background:#3d0a0a;border:1px solid #f85149;padding:14px 18px;
    border-radius:8px;color:#ffa198;margin:8px 0}
.risk-banner-HIGH{background:#2d1a00;border:1px solid #e3b341;padding:14px 18px;
    border-radius:8px;color:#f0c674;margin:8px 0}
.risk-banner-MEDIUM{background:#1a2200;border:1px solid #7ee787;padding:14px 18px;
    border-radius:8px;color:#9be89e;margin:8px 0}
.risk-banner-LOW{background:#0d1f2d;border:1px solid #38bdf8;padding:14px 18px;
    border-radius:8px;color:#7dd3fc;margin:8px 0}
.impact-hit{background:#2d1a00;border-left:4px solid #e3b341;padding:9px 12px;
    border-radius:5px;font-size:13px;margin:4px 0;color:#f0c674}
.impact-none{background:#0f2a1e;border-left:4px solid #3fb950;padding:9px 12px;
    border-radius:5px;font-size:13px;color:#7ee787}
.dead-card{background:#3d0a0a;border-left:4px solid #f85149;padding:8px 12px;
    border-radius:5px;font-size:12px;margin:3px 0;color:#ffa198}
.compare-box{background:#161b22;border:1px solid #30363d;border-radius:8px;
    padding:14px;font-size:13px;color:#c9d1d9;margin:4px 0}
.compare-shared{background:#0f2a1e;border-left:3px solid #3fb950;padding:5px 10px;
    border-radius:4px;font-size:12px;margin:2px 0;color:#7ee787}
.compare-only{background:#162032;border-left:3px solid #38bdf8;padding:5px 10px;
    border-radius:4px;font-size:12px;margin:2px 0;color:#7dd3fc}
.chat-user{background:#1f2937;border-radius:12px 12px 2px 12px;padding:10px 14px;
    margin:5px 0;font-size:13px;color:#c9d1d9}
.chat-ai{background:#1e1120;border-radius:2px 12px 12px 12px;padding:10px 14px;
    margin:5px 0;font-size:13px;color:#c9d1d9;line-height:1.6}
.tool-card{background:#0d1117;border:1px solid #30363d;border-radius:10px;
    padding:16px;margin-bottom:14px}
</style>
""", unsafe_allow_html=True)

# =====================================================
# CONFIG
# =====================================================

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
META_FOLDER = os.path.join(BASE_DIR, "Meta_Data_File")

# =====================================================
# METADATA LOADERS
# =====================================================

def _read_meta(path):
    if not path:
        return None
    try:
        if path.lower().endswith(".csv"):
            for enc in ["utf-8","utf-8-sig","cp1252","latin1","ISO-8859-1"]:
                try:
                    return pd.read_csv(path, encoding=enc, engine="python", on_bad_lines="skip")
                except:
                    pass
        elif path.lower().endswith((".xls",".xlsx")):
            return pd.read_excel(path)
    except:
        pass
    return None


def _find(pattern):
    files = []
    for ext in [".csv",".xls",".xlsx"]:
        files.extend(glob.glob(os.path.join(META_FOLDER, pattern + ext)))
    return files[0] if files else None


def _filter_active(df):
    if df is None:
        return None
    df = df.astype(str)
    if "OBJVERS" in df.columns:
        df = df[df["OBJVERS"].str.strip().str.upper() == "A"].copy()
    return df


@st.cache_data(show_spinner=False)
def load_metadata():
    raw_keys = {
        "query_df": "*Query*",
        "mp_df":    "*MultiProvider*",
        "trfn_df":  "*TRANSFORMATION*",
        "dtp_df":   "*DTP*",
        "cube_df":  "*Infocube*",
        "adso_df":  "*ADSO*",
        "ip_df":    "*Infopackage*",
    }
    dfs = {}
    for key, pattern in raw_keys.items():
        df = _filter_active(_read_meta(_find(pattern)))
        if df is not None and len(df) > 0:
            dfs[key] = df

    if dfs.get("trfn_df") is not None:
        for col in ["SOURCENAME","TARGETNAME"]:
            if col in dfs["trfn_df"].columns:
                dfs["trfn_df"][col] = (
                    dfs["trfn_df"][col]
                    .str.replace(r"GDVCLNT\d+","",regex=True).str.strip()
                )
    if dfs.get("dtp_df") is not None:
        for col in ["SRC","TGT"]:
            if col in dfs["dtp_df"].columns:
                dfs["dtp_df"][col] = (
                    dfs["dtp_df"][col]
                    .str.replace(r"GDVCLNT\d+","",regex=True).str.strip()
                )

    dfs["_mp_set"]   = set(dfs["mp_df"]["MULTIPROVIDER"].str.upper())  if dfs.get("mp_df")   is not None else set()
    dfs["_cube_set"] = set(dfs["cube_df"]["INFOCUBE"].str.upper())     if dfs.get("cube_df") is not None else set()
    dfs["_adso_set"] = set(dfs["adso_df"]["ADSO"].str.upper())         if dfs.get("adso_df") is not None else set()
    return dfs

# =====================================================
# OBJECT TYPE
# =====================================================

def get_object_type(obj, dfs):
    obj = str(obj).upper().strip()
    if obj in dfs["_mp_set"]:   return "MULTIPROVIDER"
    if obj in dfs["_cube_set"]: return "INFOCUBE"
    if obj in dfs["_adso_set"]: return "ADSO"
    if obj.endswith("_TR"):     return "INFOSOURCE"
    return "DATASOURCE"

# =====================================================
# BUILD LINEAGE
# =====================================================

def build_lineage(provider, dfs):
    visited = set()
    stats   = {"transformations": set(), "dtps": set(), "ip_count": 0}
    trfn_df = dfs.get("trfn_df")
    dtp_df  = dfs.get("dtp_df")
    mp_df   = dfs.get("mp_df")
    ip_df   = dfs.get("ip_df")

    def _proc(prov):
        pu = prov.upper()
        if pu in visited: return None
        visited.add(pu)
        node = {
            "provider": prov, "type": get_object_type(prov, dfs),
            "children": [], "dtps": [], "transformations": [],
            "ip_nodes": [], "infopackages": 0,
        }
        if dtp_df is not None and "SRC" in dtp_df.columns and "TGT" in dtp_df.columns:
            base = prov.replace("_TR","")
            dm = dtp_df[dtp_df["SRC"].str.upper().eq(base.upper()) & dtp_df["TGT"].str.upper().eq(pu)]
            if not dm.empty:
                dtps = dm["DTP"].dropna().astype(str).tolist()
                node["dtps"] = dtps; stats["dtps"].update(dtps)
        if trfn_df is not None and "TARGETNAME" in trfn_df.columns:
            tm = trfn_df[trfn_df["TARGETNAME"].str.upper().eq(pu)]
            if not tm.empty:
                trs = tm["TRANID"].dropna().astype(str).tolist()
                node["transformations"] = trs; stats["transformations"].update(trs)
        if ip_df is not None and "OLTPSOURCE" in ip_df.columns:
            base = prov.replace("_TR","")
            ip_rows = ip_df[ip_df["OLTPSOURCE"].str.upper().eq(base.upper())]
            node["infopackages"] = len(ip_rows); stats["ip_count"] += len(ip_rows)
            for _, irow in ip_rows.iterrows():
                ip_id = str(irow.get("LOGDPID","")).strip()
                if ip_id and ip_id.lower() != "nan":
                    node["ip_nodes"].append(ip_id)
        if node["type"] == "MULTIPROVIDER" and mp_df is not None:
            for cube in mp_df[mp_df["MULTIPROVIDER"].str.upper().eq(pu)]["PARTPROVIDER"].dropna().unique():
                ch = _proc(str(cube).strip())
                if ch: node["children"].append(ch)
        if trfn_df is not None and "TARGETNAME" in trfn_df.columns:
            for _, row in trfn_df[trfn_df["TARGETNAME"].str.upper().eq(pu)].iterrows():
                tranid = str(row["TRANID"]).strip()
                source = str(row["SOURCENAME"]).strip()
                base_s = source.replace("_TR","")
                stats["transformations"].add(tranid)
                dtp_list = []
                if dtp_df is not None and "SRC" in dtp_df.columns and "TGT" in dtp_df.columns:
                    dm2 = dtp_df[dtp_df["SRC"].str.upper().eq(base_s.upper()) & dtp_df["TGT"].str.upper().eq(pu)]
                    if not dm2.empty:
                        dtp_list = dm2["DTP"].dropna().astype(str).tolist()
                        stats["dtps"].update(dtp_list)
                ip_list = []
                if ip_df is not None and "OLTPSOURCE" in ip_df.columns:
                    ipm2 = ip_df[ip_df["OLTPSOURCE"].str.upper().eq(base_s.upper())]
                    stats["ip_count"] += len(ipm2)
                    for _, irow in ipm2.iterrows():
                        ip_id = str(irow.get("LOGDPID","")).strip()
                        if ip_id and ip_id.lower() != "nan":
                            ip_list.append(ip_id)
                node["children"].append({
                    "transformation": tranid, "source": source,
                    "lineage": _proc(source),
                    "dtp_list": dtp_list, "dtps": len(dtp_list),
                    "ip_list": ip_list,   "infopackages": len(ip_list),
                })
        return node

    return _proc(str(provider).strip()), stats

# =====================================================
# COMPLEXITY
# =====================================================

def compute_complexity(stats):
    t = len(stats["transformations"])
    d = len(stats["dtps"])
    i = stats["ip_count"]
    score = (t * 5) + (d * 3) + i
    cx = "LOW" if score < 10 else ("MEDIUM" if score < 30 else "HIGH")
    return t, d, i, score, cx

# =====================================================
# FLATTEN TREE
# =====================================================

def flatten_tree(node, rows=None, depth=0):
    if rows is None: rows = []
    if not node: return rows
    rows.append({"type": node.get("type","DATASOURCE"), "name": node["provider"], "depth": depth})
    for tr in node.get("transformations",[]): rows.append({"type":"TRANSFORMATION","name":tr,"depth":depth+1})
    for dtp in node.get("dtps",[]): rows.append({"type":"DTP","name":dtp,"depth":depth+1})
    for ip in node.get("ip_nodes",[]): rows.append({"type":"INFOPACKAGE","name":ip,"depth":depth+1})
    for child in node.get("children",[]):
        if "transformation" in child:
            rows.append({"type":"TRANSFORMATION","name":child["transformation"],"depth":depth+1})
            rows.append({"type":"SOURCE","name":child["source"],"depth":depth+2})
            for dtp in child.get("dtp_list",[]): rows.append({"type":"DTP","name":dtp,"depth":depth+3})
            for ip in child.get("ip_list",[]): rows.append({"type":"INFOPACKAGE","name":ip,"depth":depth+3})
            if child.get("lineage"): flatten_tree(child["lineage"],rows,depth+2)
        else:
            flatten_tree(child,rows,depth+1)
    return rows

# =====================================================
# VISUAL RENDERER
# =====================================================

_TYPE_META = {
    "QUERY":          ("ln-query",       "🔍", "BEx Query"),
    "INFOPROVIDER":   ("ln-provider",    "🔀", "InfoProvider"),
    "MULTIPROVIDER":  ("ln-mp",          "🔀", "MultiProvider"),
    "INFOCUBE":       ("ln-cube",        "🧊", "InfoCube"),
    "ADSO":           ("ln-adso",        "🗄️", "ADSO"),
    "TRANSFORMATION": ("ln-trfn",        "⚙️", "Transformation"),
    "DTP":            ("ln-dtp",         "🔁", "DTP"),
    "DATASOURCE":     ("ln-datasource",  "📡", "DataSource"),
    "SOURCE":         ("ln-datasource",  "📡", "DataSource"),
    "INFOSOURCE":     ("ln-infosource",  "📥", "InfoSource"),
    "INFOPACKAGE":    ("ln-infopackage", "📬", "InfoPackage"),
    "SOURCE SYSTEM":  ("ln-srcsys",      "🏭", "Source System"),
}

def render_lineage_visual(rows):
    if not rows: return
    html = []
    for row in rows:
        otype  = str(row["type"]).upper()
        depth  = min(row["depth"], 12)
        indent = depth * 18
        cls, icon, label = _TYPE_META.get(otype, ("ln-default","📦",otype))
        connector = "└─" if depth > 0 else ""
        html.append(
            f"<div class='ln-row {cls}' style='margin-left:{indent}px'>"
            f"<span class='ln-connector'>{connector}</span>"
            f"<span>{icon}</span>"
            f"<span class='ln-type'>{label}</span>"
            f"<span style='opacity:.3'>·</span>"
            f"<span class='ln-name'>{row['name']}</span>"
            f"</div>"
        )
    st.markdown("<div class='ln-wrap'>" + "\n".join(html) + "</div>", unsafe_allow_html=True)
    seen = list(dict.fromkeys(str(r["type"]).upper() for r in rows))
    legend = []
    for otype in seen:
        cls, icon, label = _TYPE_META.get(otype, ("ln-default","📦",otype))
        legend.append(f"<span class='ln-row {cls} ln-leg-item'>{icon} {label}</span>")
    st.markdown("<div class='ln-legend'>" + "".join(legend) + "</div>", unsafe_allow_html=True)

# =====================================================
# TREE → EXCEL
# Columns: Bex Name | Root Provider | Transformation |
#          InfoSource | DTP | InfoSource Transformation |
#          DataSource | InfoPackage
# =====================================================

def tree_to_excel(query_name, provider, tree):
    """
    Build lineage Excel with merged cells:
    - Bex Name / Root Provider merged across all rows
    - Transformation / InfoSource / DTP / InfoSource Transformation / DataSource
      each merged across the rows belonging to that transformation group
    - InfoPackage gets one row per entry (the merge target)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Step 1: collect structured groups
    # Each group = one transformation with its list of InfoPackages
    # group = {trfn, infosource, dtp, is_trfn, ds, ips: [...]}
    groups = []

    def _walk_children(children):
        for child in children:
            if "transformation" not in child:
                _walk_children(child.get("children", [])); continue

            trfn     = child.get("transformation", "")
            source   = child.get("source", "")
            dtp_list = child.get("dtp_list", [])
            ip_list  = child.get("ip_list", [])

            is_trfn = ""; datasource = ""
            lineage = child.get("lineage")
            if lineage:
                for lchild in lineage.get("children", []):
                    if "transformation" in lchild:
                        is_trfn    = lchild.get("transformation", "")
                        datasource = lchild.get("source", "")
                        if lchild.get("lineage"):
                            for sub in lchild["lineage"].get("children", []):
                                if "transformation" in sub:
                                    datasource = sub.get("source", datasource)
                        break

            # Each DTP beyond the first gets its own sub-group row
            all_dtps = dtp_list if dtp_list else [""]
            # We expand DTPs as separate rows within the same transformation group
            # but merge Transformation/InfoSource/IS_Transformation/DataSource across all
            ips = ip_list if ip_list else [""]
            groups.append({
                "trfn": trfn, "infosource": source,
                "dtp_list": all_dtps,
                "is_trfn": is_trfn, "ds": datasource,
                "ips": ips,
            })

    _walk_children(tree.get("children", []))

    # ── Step 2: expand groups into flat rows, tracking merge spans
    # Each flat row: {bex, prov, trfn, infosource, dtp, is_trfn, ds, ip}
    # We'll record which rows need merging per column

    flat_rows   = []  # list of value dicts
    merge_specs = []  # list of (start_row, end_row, col_indices_to_merge)
    # col indices (1-based): 1=Bex,2=Prov,3=Trfn,4=InfoSrc,5=DTP,6=ISTrfn,7=DS,8=IP

    for g in groups:
        trfn        = g["trfn"]
        infosource  = g["infosource"]
        dtp_list    = g["dtp_list"]
        is_trfn     = g["is_trfn"]
        ds          = g["ds"]
        ips         = g["ips"]

        # Total rows for this group = max(len(ips), len(dtp_list))
        # We pair DTPs and IPs: first DTP goes with first IP rows,
        # extra DTPs go as additional rows
        # Simplest approach: one row per IP, DTP shown on first row of group
        group_start = len(flat_rows)  # 0-indexed into flat_rows

        for k, ip in enumerate(ips):
            flat_rows.append({
                "trfn":      trfn       if k == 0 else "",
                "infosource":infosource if k == 0 else "",
                "dtp":       dtp_list[0] if (k == 0 and dtp_list) else "",
                "is_trfn":   is_trfn    if k == 0 else "",
                "ds":        ds         if k == 0 else "",
                "ip":        ip,
            })

        # Extra DTPs (beyond first) each get their own row below
        for ed in dtp_list[1:]:
            flat_rows.append({
                "trfn":"","infosource":"","dtp":ed,
                "is_trfn":"","ds":"","ip":"",
            })

        group_end = len(flat_rows) - 1  # 0-indexed
        span      = group_end - group_start + 1

        if span > 1:
            # Excel rows are offset by 2 (1 header + 1-indexed)
            er_start = group_start + 2
            er_end   = group_end   + 2
            # Merge cols: Transformation(3), InfoSource(4), DTP(5),
            #             InfoSource Transformation(6), DataSource(7)
            merge_specs.append((er_start, er_end, [3, 4, 5, 6, 7]))

    # Total span for Bex Name and Root Provider
    total_rows = len(flat_rows)
    if total_rows > 1:
        merge_specs.append((2, total_rows + 1, [1, 2]))

    # ── Step 3: build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lineage"

    COLS       = ["Bex Name","Root Provider","Transformation","InfoSource",
                  "DTP","InfoSource Transformation","DataSource","InfoPackage"]
    COL_WIDTHS = [22, 18, 38, 20, 38, 38, 18, 38]

    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill    = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font   = Font(name="Arial", size=9)
    # Merged cells: center vertically + horizontally
    merge_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align  = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci-1]
    ws.row_dimensions[1].height = 30

    FILLS = [
        PatternFill("solid", start_color="D6E4F7", end_color="D6E4F7"),
        PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
        PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA"),
        PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6"),
        PatternFill("solid", start_color="F4E1F7", end_color="F4E1F7"),
        PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7"),
    ]

    grp_idx   = -1
    last_trfn = None

    # ── Write all cells first
    for ri, r in enumerate(flat_rows):
        excel_row = ri + 2
        if r["trfn"] and r["trfn"] != last_trfn:
            grp_idx   = (grp_idx + 1) % len(FILLS)
            last_trfn = r["trfn"]
        fill = FILLS[grp_idx] if grp_idx >= 0 else FILLS[0]

        vals = [
            query_name if ri == 0 else "",
            provider   if ri == 0 else "",
            r["trfn"], r["infosource"], r["dtp"],
            r["is_trfn"], r["ds"], r["ip"],
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=excel_row, column=ci, value=val)
            cell.font = data_font; cell.fill = fill
            cell.alignment = left_align; cell.border = border
        ws.row_dimensions[excel_row].height = 18

    # ── Apply merges
    for (er_start, er_end, cols) in merge_specs:
        if er_start >= er_end:
            continue
        for ci in cols:
            col_letter = get_column_letter(ci)
            merge_range = f"{col_letter}{er_start}:{col_letter}{er_end}"
            try:
                ws.merge_cells(merge_range)
                # Style the top-left cell (the one that shows after merge)
                cell = ws.cell(row=er_start, column=ci)
                # Determine which fill group this belongs to
                row_idx   = er_start - 2  # back to 0-indexed flat_rows
                row_trfn  = flat_rows[row_idx]["trfn"] if row_idx < len(flat_rows) else ""
                # Find the fill for this group
                gidx = -1; lt = None
                for rr in flat_rows[:row_idx+1]:
                    if rr["trfn"] and rr["trfn"] != lt:
                        gidx = (gidx + 1) % len(FILLS); lt = rr["trfn"]
                cell.fill      = FILLS[gidx] if gidx >= 0 else FILLS[0]
                cell.font      = data_font
                cell.alignment = merge_align
                cell.border    = border
            except Exception:
                pass  # skip if merge already applied or invalid range

    ws.freeze_panes = "A2"
    # Note: auto_filter disabled when cells are merged (Excel limitation)

    # ── Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ("BEx Query",       query_name),
        ("Root Provider",   provider),
        ("Total Rows",      len(flat_rows)),
        ("Transformations", len(set(r["trfn"] for r in flat_rows if r["trfn"]))),
        ("DTPs",            len(set(r["dtp"]  for r in flat_rows if r["dtp"]))),
        ("InfoPackages",    len([r for r in flat_rows if r["ip"]])),
    ]
    bold = Font(name="Arial", bold=True, size=10)
    norm = Font(name="Arial", size=10)
    for ri, (label, val) in enumerate(summary_data, 1):
        ws2.cell(row=ri, column=1, value=label).font = bold
        ws2.cell(row=ri, column=2, value=val).font   = norm
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 42

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# =====================================================
# TREE → TEXT  (kept for internal context use)
# =====================================================

def tree_to_text(tree):
    lines = []
    B, L, P = "├── ", "└── ", "│   "
    def _walk(node, level=0):
        if not node: return
        pad = "    " * level
        lines.append(f"{pad}[{node['type']}]  {node['provider']}")
        if node.get("transformations"):
            lines.append(f"{pad}{B}Transformations : {len(node['transformations'])}")
            for i,tr in enumerate(node["transformations"]):
                lines.append(f"{pad}{P}  {L if i==len(node['transformations'])-1 else B}{tr}")
        if node.get("dtps"):
            lines.append(f"{pad}{B}DTPs : {len(node['dtps'])}")
            for i,d in enumerate(node["dtps"]):
                lines.append(f"{pad}{P}  {L if i==len(node['dtps'])-1 else B}{d}")
        ip_nodes = node.get("ip_nodes",[])
        if ip_nodes:
            lines.append(f"{pad}{B}InfoPackages : {len(ip_nodes)}")
            for i,ip in enumerate(ip_nodes):
                lines.append(f"{pad}{P}  {L if i==len(ip_nodes)-1 else B}{ip}")
        for child in node.get("children",[]):
            if "transformation" in child:
                lines.append(f"{pad}{B}Transformation : {child['transformation']}")
                lines.append(f"{pad}{P}  {B}Source       : {child['source']}")
                dl = child.get("dtp_list",[])
                lines.append(f"{pad}{P}  {B}DTPs : {len(dl)}")
                for d in dl: lines.append(f"{pad}{P}  {P}  {B}{d}")
                il = child.get("ip_list",[])
                lines.append(f"{pad}{P}  {L}InfoPackages : {len(il)}")
                for i,ip in enumerate(il):
                    lines.append(f"{pad}{P}      {L if i==len(il)-1 else B}{ip}")
                if child.get("lineage"): _walk(child["lineage"],level+1)
            else:
                _walk(child,level+1)
    _walk(tree)
    return "\n".join(lines)

# =====================================================
# API KEY
# =====================================================

def get_api_key():
    try:
        k = st.secrets.get("ANTHROPIC_API_KEY","")
        if k: return k
    except: pass
    return st.session_state.get("api_key","").strip()

# =====================================================
# AI CALLERS
# =====================================================

import requests, json, re as _re

def call_claude(system, user, max_tokens=900):
    key = get_api_key()
    if not key: return "[ERROR: No API key — paste it in the sidebar]"
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":key,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-6","max_tokens":max_tokens,"system":system,
                  "messages":[{"role":"user","content":user}]},
            timeout=60
        )
        if r.status_code != 200: return f"[ERROR {r.status_code}: {r.text[:400]}]"
        return "".join(b.get("text","") for b in r.json().get("content",[]))
    except Exception as e: return f"[ERROR: {e}]"

def call_claude_chat(system, messages, max_tokens=600):
    key = get_api_key()
    if not key: return "[ERROR: No API key — paste it in the sidebar]"
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":key,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-6","max_tokens":max_tokens,"system":system,"messages":messages},
            timeout=60
        )
        if r.status_code != 200: return f"[ERROR {r.status_code}: {r.text[:400]}]"
        return "".join(b.get("text","") for b in r.json().get("content",[]))
    except Exception as e: return f"[ERROR: {e}]"

def make_lineage_context(query_input, provider, tree, stats):
    t,d,i,score,cx = compute_complexity(stats)
    lines = tree_to_text(tree).splitlines()[:150]
    return (f"BEx Query: {query_input}\nProvider: {provider}\n"
            f"Complexity: {cx} (Score:{score})\n"
            f"Transformations:{t} | DTPs:{d} | InfoPackages:{i}\n\nLineage Tree:\n" + "\n".join(lines))

def do_ai_summary(query_input, provider, tree, stats):
    ctx = make_lineage_context(query_input, provider, tree, stats)
    return call_claude(
        "You are an SAP BW/4HANA expert. Explain lineage trees clearly.",
        f"Write a plain-English summary (5-8 sentences) covering: what this query "
        f"reports on, the data flow path, key objects, and what drives the complexity.\n\n{ctx}",
        max_tokens=600
    )

def do_risk_analysis(query_input, provider, tree, stats):
    ctx = make_lineage_context(query_input, provider, tree, stats)
    prompt = (
        "Analyze this SAP BW lineage for BW/4HANA migration risks.\n"
        "Reply with ONLY a JSON object — no markdown, no prose, no code fences.\n"
        'Structure: {"overall_risk":"HIGH","summary":"one sentence",'
        '"risks":[{"level":"HIGH","object":"obj","issue":"problem","recommendation":"fix"}],'
        '"migration_effort_days":10,"key_actions":["a1","a2"]}\n'
        f"Levels: CRITICAL, HIGH, MEDIUM, LOW only. migration_effort_days: integer.\n\nLineage:\n{ctx}"
    )
    raw = call_claude("You are an SAP BW to BW/4HANA migration expert. Respond with only valid JSON.",
                      prompt, max_tokens=2000)
    if raw.startswith("[ERROR"): return {"_error": raw}
    clean = raw.strip()
    if "```" in clean:
        clean = _re.sub(r"```[a-z]*","",clean).replace("```","").strip()
    try: return json.loads(clean)
    except: pass
    m = _re.search(r'\{[\s\S]*\}', clean)
    if m:
        try: return json.loads(m.group())
        except: pass
    try:
        trimmed  = clean.rstrip()
        repaired = trimmed + "]"*max(0,trimmed.count("[")-trimmed.count("]"))
        repaired = repaired + "}"*max(0,repaired.count("{")-repaired.count("}"))
        return json.loads(repaired)
    except: pass
    return {"_error": f"Could not parse JSON.\n\nRaw:\n{clean[:600]}"}

def do_chat(question, ctx, history):
    system = "You are an SAP BW/4HANA lineage expert. Answer questions about the lineage concisely and accurately."
    messages = []
    for h in history:
        messages.append({"role":"user","content":h["user"]})
        messages.append({"role":"assistant","content":h["ai"]})
    messages.append({"role":"user","content":f"Lineage Context:\n{ctx}\n\nQuestion: {question}"})
    return call_claude_chat(system, messages, max_tokens=600)

def collect_all_objects(dfs):
    query_df = dfs.get("query_df"); trfn_df = dfs.get("trfn_df"); obj_map = {}
    if query_df is None: return obj_map
    for _, row in query_df.iterrows():
        qid = str(row.get("QUERYID","")).strip(); provider = str(row.get("INFOPROVIDER","")).strip()
        if not qid or not provider: continue
        visited = set()
        def _collect(prov):
            pu = prov.upper()
            if pu in visited: return
            visited.add(pu); obj_map.setdefault(pu, set()).add(qid)
            if trfn_df is not None and "TARGETNAME" in trfn_df.columns:
                for _, r in trfn_df[trfn_df["TARGETNAME"].str.upper().eq(pu)].iterrows():
                    src = str(r["SOURCENAME"]).strip()
                    obj_map.setdefault(src.upper(), set()).add(qid); _collect(src)
            if dfs.get("mp_df") is not None and "PARTPROVIDER" in dfs["mp_df"].columns:
                for part in dfs["mp_df"][dfs["mp_df"]["MULTIPROVIDER"].str.upper().eq(pu)]["PARTPROVIDER"].dropna():
                    _collect(str(part).strip())
        _collect(provider)
    return obj_map

def do_impact_analysis(obj_name, dfs):
    obj_map  = collect_all_objects(dfs)
    key      = obj_name.upper().strip()
    impacted = sorted(obj_map.get(key, set()))
    partial  = {obj: sorted(qs) for obj,qs in obj_map.items() if key in obj and obj != key}
    ctx = (f"Object: {obj_name}\nDirect hits: {len(impacted)}\n"
           f"Queries: {', '.join(impacted[:30]) or 'None'}\nPartial matches: {list(partial.keys())[:10]}")
    ai = call_claude("You are an SAP BW/4HANA expert.",
                     f"Explain impact of changing/removing this object.\n\n{ctx}\n\n"
                     f"3-5 sentences: what it likely is, which queries break, precautions to take.",
                     max_tokens=400)
    return {"object":obj_name,"impacted":impacted,"partial":partial,"ai":ai}

def do_dead_objects(dfs):
    active = set(collect_all_objects(dfs).keys()); dead = []; seen = set()
    checks = [("trfn_df","TARGETNAME","TRANSFORMATION"),("dtp_df","DTP","DTP"),
              ("cube_df","INFOCUBE","INFOCUBE"),("adso_df","ADSO","ADSO"),
              ("ip_df","OLTPSOURCE","DATASOURCE/INFOPACKAGE")]
    for df_key,col,label in checks:
        df = dfs.get(df_key)
        if df is not None and col in df.columns:
            for obj in df[col].dropna().str.upper().unique():
                obj = obj.strip()
                if obj and obj not in active and obj not in seen:
                    seen.add(obj); dead.append({"type":label,"object":obj})
    counts = dict(Counter(d["type"] for d in dead))
    ctx = (f"Dead objects: {len(dead)}\nBreakdown: {counts}\n"
           f"Active: {len(active)}\nSample: {[d['object'] for d in dead[:15]]}")
    ai = call_claude("You are an SAP BW/4HANA housekeeping expert.",
                     f"Explain these dead objects.\n\n{ctx}\n\n"
                     f"4-5 sentences: why they accumulate, deletion risk, recommended steps.",
                     max_tokens=400)
    return {"dead":dead,"counts":counts,"active_count":len(active),"ai":ai}

def do_query_comparison(q1, q2, dfs):
    def _get(q):
        qdf = dfs.get("query_df")
        if qdf is None: return None, set(), None, None
        m = qdf[qdf["QUERYID"].str.contains(q,case=False,na=False)|qdf["QUERYNAME"].str.contains(q,case=False,na=False)] \
            if "QUERYID" in qdf.columns and "QUERYNAME" in qdf.columns else pd.DataFrame()
        if m.empty: return None, set(), None, None
        prov = str(m.iloc[0]["INFOPROVIDER"]).strip()
        tree, stats = build_lineage(prov, dfs)
        objs = set()
        def _walk(node):
            if not node: return
            objs.add(node.get("provider","").upper())
            for ch in node.get("children",[]):
                if "transformation" in ch:
                    objs.add(ch.get("source","").upper()); objs.add(ch.get("transformation","").upper())
                    _walk(ch.get("lineage"))
                else: _walk(ch)
        _walk(tree)
        return prov, objs, tree, stats
    p1,o1,t1,s1 = _get(q1); p2,o2,t2,s2 = _get(q2)
    if p1 is None: return {"error":f"Query '{q1}' not found"}
    if p2 is None: return {"error":f"Query '{q2}' not found"}
    shared=sorted(o1&o2); only_q1=sorted(o1-o2); only_q2=sorted(o2-o1)
    st1=compute_complexity(s1); st2=compute_complexity(s2)
    ctx=(f"Q1:{q1} P:{p1} Cx:{st1[4]} Score:{st1[3]} T:{st1[0]} D:{st1[1]} IP:{st1[2]}\n"
         f"Q2:{q2} P:{p2} Cx:{st2[4]} Score:{st2[3]} T:{st2[0]} D:{st2[1]} IP:{st2[2]}\n"
         f"Shared({len(shared)}):{', '.join(shared[:20])}\n"
         f"Only Q1({len(only_q1)}):{', '.join(only_q1[:15])}\n"
         f"Only Q2({len(only_q2)}):{', '.join(only_q2[:15])}")
    ai = call_claude("You are an SAP BW/4HANA migration expert.",
                     f"Compare these two queries (5-7 sentences): similarity, shared object risks, "
                     f"consolidation opportunities, which is harder to migrate.\n\n{ctx}",
                     max_tokens=500)
    return {"q1":q1,"q2":q2,"p1":p1,"p2":p2,"st1":st1,"st2":st2,
            "shared":shared,"only_q1":only_q1,"only_q2":only_q2,"ai":ai}

# =====================================================
# SESSION STATE
# =====================================================

for _k, _v in [
    ("result",None),("data",None),("last_file",None),
    ("lin_tree",None),("lin_stats",None),
    ("lin_query",None),("lin_error",None),
    ("ai_sum",None),("ai_risk",None),
    ("chat_hist",[]),
    ("ia_result",None),("dead_result",None),("cmp_result",None),
    ("lin_provider",""),("lin_results",None),
    ("trfn_results",[]),("trfn_processing",False),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# =====================================================
# HEADER
# =====================================================

st.markdown("<div class='hero-title'>⚙️ BW Transformation Analyzer</div>", unsafe_allow_html=True)
st.markdown("<div class='hero-sub'>AI-powered BW Transformation Analysis & Full Lineage Explorer</div>", unsafe_allow_html=True)
st.markdown("---")

dfs = load_metadata()

# =====================================================
# LAYOUT
# =====================================================

left_col, right_col = st.columns([1,1], gap="large")

# ─────────────────────────────────────────
# LEFT — Transformation Analyzer
# ─────────────────────────────────────────
with left_col:
    st.markdown("#### 📂 Transformation Analyzer")
    lt1, lt2 = st.tabs(["📁 Single File", "📋 Batch via CSV / Excel"])
    uploaded_file = None
    batch_files   = []

    with lt1:
        st.caption("Upload one Excel file exported from your SAP BW transformation.")
        uploaded_file = st.file_uploader("Excel file", type=["xlsx","xls"],
                                         key="single_trfn", label_visibility="collapsed")

    with lt2:
        st.caption("Upload multiple transformation Excel files at once.")
        batch_files = st.file_uploader("Transformation files (multi-select)", type=["xlsx","xls"],
                                       accept_multiple_files=True, key="batch_trfn",
                                       label_visibility="collapsed")
        if batch_files:
            st.success(f"✅ {len(batch_files)} file(s) selected")

    def _analyze_one(uf):
        try:
            uf.seek(0)
            d = process_transformation(uf)
            r = analyze_code(d["code"], d["summary"])
            return {"file":uf.name,"data":d,"result":r,"error":None}
        except Exception as e:
            return {"file":uf.name,"data":None,"result":None,"error":str(e)}

    if uploaded_file:
        if uploaded_file.name != st.session_state.last_file:
            st.session_state.last_file = uploaded_file.name
            st.session_state.result    = None
            st.session_state.data      = None
            st.session_state.trfn_results = []
        if st.session_state.result is None:
            with st.spinner("🤖 Analyzing transformation…"):
                out = _analyze_one(uploaded_file)
            if out["error"]: st.error(f"❌ {out['error']}")
            else:
                st.session_state.data   = out["data"]
                st.session_state.result = out["result"]

    if batch_files:
        if st.button("▶️ Analyze All Files", type="primary", key="batch_run"):
            st.session_state.trfn_results = []
            prog = st.progress(0, text="Analyzing batch…")
            for i, uf in enumerate(batch_files):
                prog.progress(i/len(batch_files), text=f"Analyzing: {uf.name}")
                st.session_state.trfn_results.append(_analyze_one(uf))
            prog.progress(1.0, text="Done!")

    result = st.session_state.result
    data   = st.session_state.data

    if result and not st.session_state.trfn_results:
        cx     = str(result.get("complexity","UNKNOWN")).upper()
        cx_cls = {"LOW":"cx-LOW","MEDIUM":"cx-MEDIUM","HIGH":"cx-HIGH"}.get(cx,"")
        _, mid, _ = st.columns([1,1,1])
        with mid:
            st.markdown(f"<div class='metric-card'><h3>Complexity</h3><h2 class='{cx_cls}'>{cx}</h2></div>",
                        unsafe_allow_html=True)
        st.markdown("")
        m1,m2,m3 = st.columns(3)
        m1.metric("📄 Code Lines", result.get("code_lines","—"))
        m2.metric("🗂️ Z-Tables",   len(result.get("ztables",[])) if result.get("ztables") else 0)
        m3.metric("🧠 Confidence", result.get("confidence","—"))
        st.markdown("")
        t1,t2,t3,t4 = st.tabs(["📋 Overview","🤖 AI Analysis","🗂️ Z-Tables","📊 Summary Table"])
        with t1:
            st.markdown(f"<div class='ai-box'>{result.get('overview','No overview.')}</div>", unsafe_allow_html=True)
        with t2:
            st.markdown(result.get("reason","No analysis returned."))
            if st.button("🔄 Regenerate", key="regen"):
                st.session_state.result = None; st.rerun()
        with t3:
            if result.get("ztables"):
                st.info(result.get("ztable_summary",""))
                for zt in result["ztables"]:
                    st.markdown(f"<div class='ztable-card'>📦 <b>{zt}</b></div>", unsafe_allow_html=True)
            else:
                st.success("✅ No custom Z-Tables detected.")
        with t4:
            if data and data.get("summary") is not None:
                st.dataframe(data["summary"], use_container_width=True)
            else:
                st.info("No summary available.")
        st.markdown("---")
        report_df = pd.DataFrame([{
            "File":            uploaded_file.name if uploaded_file else "",
            "Complexity":      result.get("complexity",""),
            "Overview":        result.get("overview",""),
            "AI Analysis":     result.get("reason",""),
            "Z-Tables":        ", ".join(result.get("ztables",[])) if result.get("ztables") else "",
            "Z-Table Summary": result.get("ztable_summary",""),
        }])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            report_df.to_excel(writer, index=False, sheet_name="Analysis")
            if data and data.get("summary") is not None:
                data["summary"].to_excel(writer, index=False, sheet_name="Transformation Summary")
        buf.seek(0)
        st.download_button(
            "📥 Download Report (.xlsx)", data=buf,
            file_name=f"bw_analysis_{uploaded_file.name if uploaded_file else 'report'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if st.session_state.trfn_results:
        st.markdown(f"#### Batch Results — {len(st.session_state.trfn_results)} files")
        summary_rows = []
        for r in st.session_state.trfn_results:
            if r["error"]:
                summary_rows.append({"File":r["file"],"Complexity":"ERROR","Z-Tables":"","Error":r["error"]})
            else:
                res = r["result"]
                summary_rows.append({"File":r["file"],"Complexity":res.get("complexity","—"),
                                     "Z-Tables":len(res.get("ztables",[])) if res.get("ztables") else 0,"Error":""})
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)
        for r in st.session_state.trfn_results:
            with st.expander(f"{'❌' if r['error'] else '✅'} {r['file']}"):
                if r["error"]: st.error(r["error"])
                else:
                    res = r["result"]
                    st.markdown(f"**Complexity:** `{str(res.get('complexity','?')).upper()}`")
                    st.markdown(f"<div class='ai-box'>{res.get('overview','')}</div>", unsafe_allow_html=True)
                    if res.get("ztables"):
                        st.markdown(f"**Z-Tables:** {', '.join(res['ztables'])}")
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf,"w",zipfile.ZIP_DEFLATED) as zf:
            for r in st.session_state.trfn_results:
                if r["result"]:
                    res = r["result"]
                    df  = pd.DataFrame([{"File":r["file"],"Complexity":res.get("complexity",""),
                                         "Overview":res.get("overview",""),"AI Analysis":res.get("reason",""),
                                         "Z-Tables":", ".join(res.get("ztables",[])) if res.get("ztables") else ""}])
                    xbuf = io.BytesIO()
                    with pd.ExcelWriter(xbuf, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Analysis")
                        if r["data"] and r["data"].get("summary") is not None:
                            r["data"]["summary"].to_excel(writer, index=False, sheet_name="Summary")
                    safe = "".join(c if c.isalnum() or c in "-_." else "_" for c in r["file"])
                    zf.writestr(f"analysis_{safe}.xlsx", xbuf.getvalue())
        zip_buf.seek(0)
        st.download_button(f"📥 Download All {len(st.session_state.trfn_results)} Reports (.zip)",
                           data=zip_buf, file_name="bw_batch_analysis.zip",
                           mime="application/zip", key="batch_dl")

# ─────────────────────────────────────────
# RIGHT — BW Lineage Explorer
# ─────────────────────────────────────────
with right_col:
    st.markdown("#### 🔗 BW Lineage Explorer")
    st.caption("Full chain — Query → InfoCube/ADSO → Transformation → InfoSource → DTP → DataSource → InfoPackages")

    loaded = sum(1 for k,v in dfs.items() if not k.startswith("_") and v is not None)
    if loaded == 7:   st.success("✅ All 7 metadata files loaded (OBJVERS = A only)")
    elif loaded > 0:  st.warning(f"⚠️ {loaded}/7 metadata files loaded")
    else:             st.error("❌ No metadata files found in Meta_Data_File folder")

    st.markdown("---")
    st.markdown("### 🔑 API Key")
    try:    has_secret = bool(st.secrets.get("ANTHROPIC_API_KEY",""))
    except: has_secret = False
    if has_secret:
        st.success("✅ API key loaded from secrets")
    else:
        st.text_input("Anthropic API key", type="password", placeholder="sk-ant-...",
                      key="api_key", help="Get your key at console.anthropic.com")
        if st.session_state.get("api_key",""):  st.success("✅ Key set for this session")
        else:                                    st.warning("⚠️ No key — AI features disabled")

    st.markdown("---")
    st.markdown("### 🤖 AI Features")
    ai_sum_on  = st.toggle("✨ AI Summary",              value=True, key="t_sum")
    ai_risk_on = st.toggle("🛡️ Migration Risk Analysis", value=True, key="t_risk")
    ai_chat_on = st.toggle("💬 Chat with Lineage",       value=True, key="t_chat")
    ai_ia_on   = st.toggle("🔎 Impact Analysis",         value=True, key="t_ia")
    ai_dead_on = st.toggle("🧹 Dead Object Detector",    value=True, key="t_dead")
    ai_cmp_on  = st.toggle("⚖️ Query Comparison",        value=True, key="t_cmp")

    inp_tab1, inp_tab2 = st.tabs(["✏️ Single Query", "📄 Upload CSV / Excel"])
    queries_to_run = []

    with inp_tab1:
        st.caption("Enter one query name or ID (partial names work — uses contains match).")
        single_q = st.text_input("Query name or ID", placeholder="e.g. 0IC_C03_Q0001",
                                 key="single_q", label_visibility="collapsed")
        if single_q.strip(): queries_to_run = [single_q.strip()]

    with inp_tab2:
        st.caption("Upload a CSV or Excel file with a QUERYID or QUERYNAME column.")
        qfile = st.file_uploader("Query list file", type=["csv","xls","xlsx"],
                                 key="qfile_upload", label_visibility="collapsed")
        if qfile:
            try:
                qfile.seek(0)
                if qfile.name.lower().endswith(".csv"):
                    qf_df = None
                    for enc in ["utf-8","utf-8-sig","cp1252","latin1"]:
                        try:
                            qfile.seek(0)
                            qf_df = pd.read_csv(qfile, encoding=enc, engine="python", on_bad_lines="skip"); break
                        except: pass
                else:
                    qf_df = pd.read_excel(qfile)
                if qf_df is not None and len(qf_df) > 0:
                    qf_df = qf_df.astype(str)
                    if   "QUERYID"   in qf_df.columns: raw_qs = qf_df["QUERYID"].tolist()
                    elif "QUERYNAME" in qf_df.columns: raw_qs = qf_df["QUERYNAME"].tolist()
                    else:                               raw_qs = qf_df.iloc[:,0].tolist()
                    queries_to_run = [q.strip() for q in raw_qs if q.strip() and q.strip().lower() != "nan"]
                    st.success(f"✅ {len(queries_to_run)} quer{'y' if len(queries_to_run)==1 else 'ies'} loaded")
                    with st.expander(f"Preview ({min(10,len(queries_to_run))} shown)"):
                        for q in queries_to_run[:10]: st.markdown(f"- `{q}`")
                        if len(queries_to_run) > 10: st.caption(f"… and {len(queries_to_run)-10} more")
                else:
                    st.warning("File is empty or could not be read.")
            except Exception as e:
                st.error(f"Could not read file: {e}")

    b1, b2 = st.columns([1,1])
    with b1: trace = st.button("▶️ Trace Lineage", type="primary", use_container_width=True)
    with b2:
        if st.button("🗑️ Clear", use_container_width=True):
            for k in ("lin_tree","lin_stats","lin_query","lin_error","lin_provider",
                      "lin_results","ai_sum","ai_risk","chat_hist"):
                st.session_state[k] = None if k != "chat_hist" else []
            st.rerun()

    if trace and queries_to_run:
        query_df = dfs.get("query_df")
        if query_df is None:
            st.session_state.lin_error = "Query metadata not loaded."
            st.session_state.lin_results = None
        else:
            results=[]; errors=[]
            prog = st.progress(0, text="Tracing lineages…")
            for idx, q in enumerate(queries_to_run):
                prog.progress(idx/len(queries_to_run), text=f"Tracing: {q}")
                mask = pd.Series(False, index=query_df.index)
                if "QUERYID"   in query_df.columns: mask = mask | query_df["QUERYID"].str.contains(q,case=False,na=False)
                if "QUERYNAME" in query_df.columns: mask = mask | query_df["QUERYNAME"].str.contains(q,case=False,na=False)
                matches = query_df[mask]
                if matches.empty: errors.append(q); continue
                provider = str(matches.iloc[0]["INFOPROVIDER"]).strip()
                tree, stats = build_lineage(provider, dfs)
                results.append({"query":q,"provider":provider,"tree":tree,"stats":stats,
                                 "ctx":make_lineage_context(q,provider,tree,stats)})
            prog.progress(1.0, text="Done!")
            if errors: st.warning(f"⚠️ Not found: {', '.join(errors)}")
            if results:
                if len(results) == 1:
                    r = results[0]
                    st.session_state.lin_tree     = r["tree"]
                    st.session_state.lin_stats    = r["stats"]
                    st.session_state.lin_query    = r["query"]
                    st.session_state.lin_provider = r["provider"]
                    st.session_state.lin_error    = None
                    st.session_state.lin_results  = None
                else:
                    st.session_state.lin_results  = results
                    st.session_state.lin_tree     = None
                    st.session_state.lin_error    = None
                st.session_state.ai_sum    = None
                st.session_state.ai_risk   = None
                st.session_state.chat_hist = []

    # ── Render
    if st.session_state.lin_error:
        st.error(f"❌ {st.session_state.lin_error}")

    elif st.session_state.get("lin_results"):
        results = st.session_state["lin_results"]
        st.markdown(f"#### 📋 Batch Lineage — {len(results)} quer{'y' if len(results)==1 else 'ies'}")
        rows_tbl = []
        for r in results:
            t,d,i,score,cx = compute_complexity(r["stats"])
            rows_tbl.append({"Query":r["query"],"Provider":r["provider"],"Complexity":cx,
                             "Score":score,"Transforms":t,"DTPs":d,"InfoPkgs":i})
        st.dataframe(pd.DataFrame(rows_tbl), use_container_width=True)

        # Batch ZIP — Excel files
        zip_buf2 = io.BytesIO()
        with zipfile.ZipFile(zip_buf2,"w",zipfile.ZIP_DEFLATED) as zf2:
            for r in results:
                safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in r["query"])
                xbuf = tree_to_excel(r["query"], r["provider"], r["tree"])
                zf2.writestr(f"lineage_{safe}.xlsx", xbuf.getvalue())
        zip_buf2.seek(0)
        st.download_button(
            f"⬇️ Download all {len(results)} lineage reports (.zip)",
            data=zip_buf2, file_name="bw_lineage_batch.zip",
            mime="application/zip", key="batch_lin_dl"
        )
        st.markdown("")
        for r in results:
            t,d,i,score,cx = compute_complexity(r["stats"])
            with st.expander(f"🔍 {r['query']}  —  {cx}  (score {score})", expanded=False):
                mc1,mc2,mc3,mc4 = st.columns(4)
                mc1.metric("🔁 Transforms",t); mc2.metric("📦 DTPs",d)
                mc3.metric("📬 InfoPkgs",f"{i:,}"); mc4.metric("🎯 Score",score)
                render_lineage_visual(flatten_tree(r["tree"]))
                safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in r["query"])
                xbuf = tree_to_excel(r["query"], r["provider"], r["tree"])
                st.download_button("⬇️ Download (.xlsx)", data=xbuf,
                                   file_name=f"lineage_{safe}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key=f"dl_{safe}")

    elif st.session_state.lin_tree is not None:
        tree  = st.session_state.lin_tree
        stats = st.session_state.lin_stats
        lq    = st.session_state.lin_query
        prov  = str(st.session_state.get("lin_provider",""))
        t, d, i, score, cx_lin = compute_complexity(stats)
        ctx = make_lineage_context(lq, prov, tree, stats)

        st.markdown(f"**Lineage for** `{lq}`")
        mc1,mc2,mc3,mc4 = st.columns(4)
        mc1.metric("🔁 Transforms",t); mc2.metric("📦 DTPs",d)
        mc3.metric("📬 InfoPkgs",f"{i:,}"); mc4.metric("🎯 Score",score)
        st.markdown("")

        tab_lin, tab_sum, tab_risk, tab_chat = st.tabs(
            ["🔗 Lineage","✨ AI Summary","🛡️ Risk Analysis","💬 Chat"]
        )

        with tab_lin:
            render_lineage_visual(flatten_tree(tree))
            st.markdown("")
            excel_buf = tree_to_excel(lq, prov, tree)
            st.download_button("⬇️ Download Lineage (.xlsx)", data=excel_buf,
                               file_name=f"lineage_{lq}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="lin_dl")

        with tab_sum:
            if not ai_sum_on:
                st.info("Enable 'AI Summary' in the sidebar.")
            else:
                ai_s = st.session_state.get("ai_sum")
                if ai_s is None:
                    st.markdown('<div style="color:#8b949e;font-size:13px">Click below to generate an AI summary.</div>', unsafe_allow_html=True)
                    if st.button("✨ Generate AI Summary", key="gen_sum", type="primary"):
                        with st.spinner("✨ Generating…"):
                            st.session_state["ai_sum"] = do_ai_summary(lq,prov,tree,stats)
                        st.rerun()
                elif ai_s.startswith("[ERROR"):
                    st.error(ai_s)
                    if st.button("🔄 Try Again", key="retry_sum"):
                        st.session_state["ai_sum"] = None; st.rerun()
                else:
                    st.markdown(f'<div class="ai-sum-box">✨ <b>AI Summary</b><br><br>{ai_s}</div>', unsafe_allow_html=True)
                    if st.button("🔄 Regenerate", key="regen_sum"):
                        st.session_state["ai_sum"] = None; st.rerun()

        with tab_risk:
            if not ai_risk_on:
                st.info("Enable 'Migration Risk Analysis' in the sidebar.")
            else:
                rd = st.session_state.get("ai_risk")
                if rd is None:
                    st.markdown('<div style="color:#8b949e;font-size:13px">Click below to run migration risk analysis.</div>', unsafe_allow_html=True)
                    if st.button("🛡️ Run Risk Analysis", key="gen_risk", type="primary"):
                        with st.spinner("🛡️ Analyzing risks…"):
                            st.session_state["ai_risk"] = do_risk_analysis(lq,prov,tree,stats)
                        st.rerun()
                elif "_error" in rd:
                    st.error(f"⚠️ {rd['_error']}")
                    if st.button("🔄 Try Again", key="retry_risk"):
                        st.session_state["ai_risk"] = None; st.rerun()
                else:
                    overall=rd.get("overall_risk","?"); effort=rd.get("migration_effort_days","?")
                    ICONS={"CRITICAL":"🔴","HIGH":"🟠","MEDIUM":"🟡","LOW":"🟢"}
                    icon=ICONS.get(overall,"⚪")
                    rc1,rc2 = st.columns([3,1])
                    with rc1:
                        st.markdown(
                            f'<div class="risk-banner-{overall}">'
                            f'<b style="font-size:15px">{icon} Overall Risk: {overall}</b>'
                            f'<br><span style="font-size:13px;opacity:.85">{rd.get("summary","")}</span></div>',
                            unsafe_allow_html=True)
                    with rc2: st.metric("⏱️ Effort", f"{effort}d")
                    st.markdown("#### Identified Risks")
                    for r in rd.get("risks",[]):
                        lvl = r.get("level","MEDIUM")
                        st.markdown(
                            f'<div class="risk-{lvl}">'
                            f'<b>{ICONS.get(lvl,"⚪")} [{lvl}] {r.get("object","")}</b>'
                            f'<br>🔍 {r.get("issue","")}'
                            f'<br>✅ <i>{r.get("recommendation","")}</i></div>',
                            unsafe_allow_html=True)
                    acts = rd.get("key_actions",[])
                    if acts:
                        st.markdown("#### Key Actions")
                        for idx,a in enumerate(acts): st.markdown(f"**{idx+1}.** {a}")
                    if st.button("🔄 Re-analyse", key="regen_risk"):
                        st.session_state["ai_risk"] = None; st.rerun()

        with tab_chat:
            if not ai_chat_on:
                st.info("Enable 'Chat with Lineage' in the sidebar.")
            else:
                history = st.session_state.get("chat_hist",[])
                st.markdown('<div style="color:#8b949e;font-size:13px;margin-bottom:8px">💬 Ask anything about this lineage.</div>', unsafe_allow_html=True)
                for turn in history:
                    st.markdown(f'<div class="chat-user">👤 <b>You:</b> {turn["user"]}</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="chat-ai">🤖 <b>AI:</b> {turn["ai"]}</div>', unsafe_allow_html=True)
                if not history:
                    st.markdown("**Suggested questions:**")
                    suggs=[f"Summarise the data flow for {lq}","What DataSources feed this query?",
                           "How many transformation layers exist?","What should I migrate first?"]
                    sc = st.columns(2)
                    for si,sug in enumerate(suggs):
                        if sc[si%2].button(sug, key=f"sug{si}"):
                            with st.spinner("🤖 Thinking…"):
                                ans = do_chat(sug, ctx, history)
                            st.session_state["chat_hist"].append({"user":sug,"ai":ans}); st.rerun()
                user_msg = st.chat_input("Ask a question about this lineage…", key="chat_inp")
                if user_msg:
                    with st.spinner("🤖 Thinking…"):
                        ans = do_chat(user_msg, ctx, st.session_state["chat_hist"])
                    st.session_state["chat_hist"].append({"user":user_msg,"ai":ans}); st.rerun()
                if history:
                    if st.button("🗑️ Clear chat", key="clr_chat"):
                        st.session_state["chat_hist"] = []; st.rerun()

    # ── AI Tools
    st.markdown("---")
    st.markdown("#### 🛠️ AI Tools")
    tool_tab1, tool_tab2, tool_tab3 = st.tabs(["🔎 Impact Analysis","🧹 Dead Objects","⚖️ Query Comparison"])

    with tool_tab1:
        if not ai_ia_on: st.info("Enable 'Impact Analysis' in the sidebar.")
        else:
            st.markdown('<div class="tool-card">', unsafe_allow_html=True)
            st.caption("Enter any object name to find every query that depends on it.")
            ia_inp = st.text_input("Object name", placeholder="e.g. 2LIS_03_BX or 0IC_C03",
                                   key="ia_inp", label_visibility="collapsed")
            if st.button("🔎 Analyse Impact", key="ia_run", type="primary") and ia_inp.strip():
                with st.spinner(f"Scanning for '{ia_inp}'…"):
                    st.session_state["ia_result"] = do_impact_analysis(ia_inp.strip(), dfs)
            res = st.session_state.get("ia_result")
            if res:
                st.markdown(f'<div class="ai-sum-box">🔎 <b>Impact of: {res["object"]}</b><br><br>{res["ai"]}</div>', unsafe_allow_html=True)
                st.markdown(f"**{len(res['impacted'])} direct hit(s):**")
                if res["impacted"]:
                    for q in res["impacted"]: st.markdown(f'<div class="impact-hit">⚠️ <b>{q}</b></div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="impact-none">✅ No active queries depend on this object.</div>', unsafe_allow_html=True)
                if res["partial"]:
                    with st.expander(f"Partial matches ({len(res['partial'])})"):
                        for obj,qs in list(res["partial"].items())[:10]:
                            st.markdown(f"- `{obj}` — {len(qs)} quer{'y' if len(qs)==1 else 'ies'}")
                lines = [f"IMPACT: {res['object']}","="*50,f"Direct: {len(res['impacted'])}"]
                lines += [f"  {q}" for q in res["impacted"]] + ["","AI:",res["ai"]]
                ia_report = "\n".join(lines)
                st.download_button("⬇️ Download", ia_report.encode(),
                                   file_name=f"impact_{res['object']}.txt",
                                   mime="text/plain", key="ia_dl")
            st.markdown('</div>', unsafe_allow_html=True)

    with tool_tab2:
        if not ai_dead_on: st.info("Enable 'Dead Object Detector' in the sidebar.")
        else:
            st.markdown('<div class="tool-card">', unsafe_allow_html=True)
            st.caption("Scans all metadata for objects not used by any active query.")
            if st.button("🧹 Scan for Dead Objects", key="dead_run", type="primary"):
                with st.spinner("Scanning metadata…"):
                    st.session_state["dead_result"] = do_dead_objects(dfs)
            res = st.session_state.get("dead_result")
            if res:
                dc1,dc2,dc3 = st.columns(3)
                dc1.metric("💀 Dead", len(res["dead"])); dc2.metric("✅ Active", res["active_count"])
                total = len(res["dead"]) + res["active_count"]
                dc3.metric("🗑️ Cleanup %", f"{round(len(res['dead'])/total*100) if total>0 else 0}%")
                st.markdown(f'<div class="ai-sum-box">🧹 <b>AI Analysis</b><br><br>{res["ai"]}</div>', unsafe_allow_html=True)
                if res["dead"]:
                    grp = defaultdict(list)
                    for d in res["dead"]: grp[d["type"]].append(d["object"])
                    for otype, objs in sorted(grp.items()):
                        with st.expander(f"💀 {otype} — {len(objs)}"):
                            for o in objs[:80]:
                                st.markdown(f'<div class="dead-card">🗑️ <code>{o}</code></div>', unsafe_allow_html=True)
                            if len(objs)>80: st.caption(f"… and {len(objs)-80} more in download")
                else:
                    st.success("✅ No dead objects found!")
                lines = ["DEAD OBJECTS","="*50,f"Dead:{len(res['dead'])} Active:{res['active_count']}",
                         "","AI:",res["ai"],"","Objects:"]
                for d in res["dead"]: lines.append(f"[{d['type']}]  {d['object']}")
                dead_report = "\n".join(lines)
                st.download_button("⬇️ Download report", dead_report.encode(),
                                   file_name="dead_objects.txt", mime="text/plain", key="dead_dl")
            st.markdown('</div>', unsafe_allow_html=True)

    with tool_tab3:
        if not ai_cmp_on: st.info("Enable 'Query Comparison' in the sidebar.")
        else:
            st.markdown('<div class="tool-card">', unsafe_allow_html=True)
            st.caption("Compare two BEx queries — shared objects, unique paths, consolidation opportunities.")
            cca,ccb = st.columns(2)
            with cca: qc1 = st.text_input("Query 1", placeholder="Name or ID", key="qc1", label_visibility="collapsed")
            with ccb: qc2 = st.text_input("Query 2", placeholder="Name or ID", key="qc2", label_visibility="collapsed")
            if st.button("⚖️ Compare", key="cmp_run", type="primary") and qc1.strip() and qc2.strip():
                with st.spinner("Comparing…"):
                    st.session_state["cmp_result"] = do_query_comparison(qc1.strip(), qc2.strip(), dfs)
            res = st.session_state.get("cmp_result")
            if res:
                if "error" in res: st.error(res["error"])
                else:
                    t1,d1,i1,sc1,cx1=res["st1"]; t2,d2,i2,sc2,cx2=res["st2"]
                    ca,cb = st.columns(2)
                    with ca:
                        ca_html = '<div class="compare-box"><b>🔍 ' + res["q1"] + '</b><br>Provider: <code>' + res["p1"] + '</code><br>Complexity: <b>' + cx1 + '</b> (' + str(sc1) + ')<br>T:' + str(t1) + ' D:' + str(d1) + ' IP:' + str(i1) + '</div>'
                        st.markdown(ca_html, unsafe_allow_html=True)
                    with cb:
                        cb_html = '<div class="compare-box"><b>🔍 ' + res["q2"] + '</b><br>Provider: <code>' + res["p2"] + '</code><br>Complexity: <b>' + cx2 + '</b> (' + str(sc2) + ')<br>T:' + str(t2) + ' D:' + str(d2) + ' IP:' + str(i2) + '</div>'
                        st.markdown(cb_html, unsafe_allow_html=True)
                    st.markdown('<div class="ai-sum-box">⚖️ <b>AI Analysis</b><br><br>' + res["ai"] + '</div>', unsafe_allow_html=True)
                    c1,c2,c3 = st.columns(3)
                    with c1:
                        st.markdown(f"**🤝 Shared ({len(res['shared'])})**")
                        for o in res["shared"][:25]: st.markdown('<div class="compare-shared">✅ ' + o + '</div>', unsafe_allow_html=True)
                    with c2:
                        st.markdown(f"**🔵 Only in {res['q1']} ({len(res['only_q1'])})**")
                        for o in res["only_q1"][:25]: st.markdown('<div class="compare-only">· ' + o + '</div>', unsafe_allow_html=True)
                    with c3:
                        st.markdown(f"**🔵 Only in {res['q2']} ({len(res['only_q2'])})**")
                        for o in res["only_q2"][:25]: st.markdown('<div class="compare-only">· ' + o + '</div>', unsafe_allow_html=True)
                    lines=[f"COMPARISON: {res['q1']} vs {res['q2']}","="*60,
                           f"Q1: P={res['p1']} Cx={cx1}",f"Q2: P={res['p2']} Cx={cx2}",""]
                    lines+=[f"Shared ({len(res['shared'])}):"]+[f"  {o}" for o in res["shared"]]
                    lines+=[f"\nOnly {res['q1']}:"]+[f"  {o}" for o in res["only_q1"]]
                    lines+=[f"\nOnly {res['q2']}:"]+[f"  {o}" for o in res["only_q2"]]
                    lines+=["","AI:",res["ai"]]
                    safe1="".join(c if c.isalnum() else "_" for c in res["q1"])
                    safe2="".join(c if c.isalnum() else "_" for c in res["q2"])
                    cmp_txt="\n".join(lines)
                    st.download_button("⬇️ Download comparison", cmp_txt.encode(),
                                       file_name=f"compare_{safe1}_vs_{safe2}.txt",
                                       mime="text/plain", key="cmp_dl")
            st.markdown('</div>', unsafe_allow_html=True)

# =====================================================
# FOOTER
# =====================================================
st.markdown("---")
st.caption("Gyansys Made AI BW Migration Toolkit — © 2026 Gyansys. All rights reserved.")